# Project: Normalize Excel
# Description: Script to process Excel sheets by unmerging merged cells,
#              filling values, applying formatting, and saving the result.
# Author: EHCIKNA

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime
import os
import json

# ================================
# Color Definitions for Cell Fills
# ================================
# These colors can be applied to highlight issues or mark cells.
red_fill     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # Red
green_fill   = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # Green
blue_fill    = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")   # Blue
yellow_fill  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")   # Yellow
pink_fill    = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")   # Pink
# Extra optional colors
orange_fill  = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # Orange
purple_fill  = PatternFill(start_color="800080", end_color="800080", fill_type="solid")   # Purple
gray_fill    = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")   # Gray

def create_hierarchical_json(ws, json_output_file, issues, start_version="", end_version=""):
    try:
        last_row = get_last_row_with_value(ws)
        last_col = get_last_col_with_value(ws)
        # Get version row (row 2) from column 4 to last column
        version_row = [
            "" if v is None else str(v).strip()
            for v in list(ws.iter_rows(
                min_row=2, max_row=2,
                min_col=4, max_col=last_col - 1,
                values_only=True
            ))[0]
        ]
        data = {}

        # Find start column
        if start_version in (None, ""):
            start_col = 4
        else:
            if start_version in version_row:
                start_col = version_row.index(start_version) + 4  # +4 because we started from column 4
            else:
                print(f">> Start version '{start_version}' not found in row 2")
                return False

        # Find end column
        if end_version in (None, ""):
            end_col = last_col + 1  # +1 for range() exclusivity
        else:
            if end_version in version_row:
                end_col = version_row.index(end_version) + 4 + 1  # +4 for column offset, +1 for range() exclusivity
            else:
                print(f">> End version '{end_version}' not found in row 2")
                return False

        # Validate that we have columns to process
        if start_col >= end_col:
            print(">> Error: start_version must be less than end_version")
            return False

        for row in range(3, last_row + 1):
            tech      = ws.cell(row=row, column=1).value
            nodetype  = ws.cell(row=row, column=2).value
            nodename  = ws.cell(row=row, column=3).value

            # Clean and check for empty values more thoroughly
            tech = str(tech).strip() if tech is not None else ""
            nodetype = str(nodetype).strip() if nodetype is not None else ""
            nodename = str(nodename).strip() if nodename is not None else ""

            if not tech or not nodetype or not nodename:
                issues['skipped_rows_during_json_creation'].append(row)
                continue  # Skip incomplete rows

            # Extract supported nodes from columns 4 to last_col - 1
            supported_nodes = []
            for col in range(start_col, end_col):  # exclude last column
                val = ws.cell(row=row, column=col).value
                if val not in (None, ""):
                    supported_nodes.append(str(ws.cell(row=2,column=col).value))

            # Build nested structure
            data.setdefault(tech, {}).setdefault(nodetype, {})[nodename] = supported_nodes

        # Write to JSON
        with open(json_output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

        return True
    except Exception as e:
        print(f">> Function:{create_hierarchical_json.__name__}, Error: {str(e)}")
        return False


def get_last_row_with_value(ws):
    for row in range(ws.max_row,0,-1):
        if any(cell.value not in (None,"") for cell in ws[row]):
            return row
    return 0

def get_last_col_with_value(ws):
    for col in range(ws.max_column,0,-1):
        if any(cell.value not in (None,"") for cell in ws[col]):
            return col
    return 0

def remove_node_header(ws,issues):
    last_row=get_last_row_with_value(ws)
    for index in range(3,last_row+1):
        cell1=ws.cell(row=index, column=1).value
        cell2=ws.cell(row=index, column=2).value
        cell3=ws.cell(row=index, column=3).value
        if (cell1 == cell2 == cell3) or (cell2 in (None,"") and cell3 in (None,"")):
            issues['removed_header_rows'].append(index)
            ws.delete_rows(index)
    return True

def highlight_empty_cell(ws,issues):
    last_row=get_last_row_with_value(ws)
    for col in (1, 2, 3):
        for row in range(3, last_row+ 1):
            val=ws.cell(row=row, column=col).value
            if val is None or str(val).strip() == "":
                ws.cell(row=row, column=col).fill = red_fill
                issues['empty_cells_after_unmerge'].append(ws.cell(row=row, column=col).coordinate)
    last_col=get_last_col_with_value(ws)
    for row in (1, 2):
        for col in range(4, last_col+ 1):
            val=ws.cell(row=row, column=col).value
            if val is None or str(val).strip() == "":
                ws.cell(row=row, column=col).fill = red_fill
                issues['empty_cells_after_unmerge'].append(ws.cell(row=row, column=col).coordinate)
    return True

def formatting(ws):
    try:
        # Define center alignment style
        center_align = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)

        # Apply alignment to all non-empty cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = center_align
        return True
    except Exception as e:
        print(f">> Function:{formatting.__name__}, Error processing file: {str(e)}")
        return False


def unmerge_fill(ws, issues):
    try:
        # Iterate over all merged cell ranges
        for merge_range in list(ws.merged_cells.ranges):
            # Extract the top-left cell value of the merged range
            cell_value = ws[merge_range.coord.split(":")[0]].value
            if cell_value is None:
                issues['merged_empty_cells'].append(merge_range)

            # Unmerge the cell range
            ws.unmerge_cells(str(merge_range))

            # Fill each cell in the range with either value or red highlight
            for row in ws[merge_range.coord]:
                for cell in row:
                    if cell_value is None:
                        cell.fill = red_fill  # Highlight empty merged cells
                    else:
                        cell.value = cell_value  # Copy merged value

        return True
    except Exception as e:
        print(f">> Function:{unmerge_fill.__name__}, Error processing file: {str(e)}")
        return False

def validate_header_count(ws):
    try:
        last_col = get_last_col_with_value(ws)
        last_row = get_last_row_with_value(ws)
        min_required_columns = 5  # 3 fixed + 1 supported node + 1 comment
        min_required_rows = 3

        if last_col < min_required_columns:
            print(f">> Sheet has only {last_col} columns. Expected at least {min_required_columns}.")
            return False

        if last_row < min_required_rows:
            print(f">> Sheet has only {last_row} rows. Expected at least {min_required_rows}.")
            return False

        print(f">> Sheet has {last_col} columns, {last_row} rows — structure looks valid.")
        return True
    except Exception as e:
        print(f">> Function:{validate_header_count.__name__}, Error: {str(e)}")
        return False

def processing_excel(*file_info):
    input_file_name = file_info[0]
    sheet_name = file_info[1]
    output_file_name = file_info[2]
    issues = file_info[3]

    try:
        # Check input file existence
        if not os.path.exists(input_file_name):
            print(f">> '{input_file_name}' does not exist")
            return False

        # Load workbook
        wb = load_workbook(input_file_name, read_only=False)

        # Validate sheet name
        if sheet_name not in wb.sheetnames:
            print(f">> '{sheet_name}' does not exist. Available sheet names are \"{wb.sheetnames}\"")
            return False

        ws = wb[sheet_name]

        # Validation: Checking minimum rows and column in excel sheet
        validate_status=validate_header_count(ws)
        if validate_status:
            print(f">> Node Version Planner sheet validation done")
        else:
            print(f">> Node Version Planner sheet validation failed")
            return False

        # Step 1: Unmerge merged cells
        unmerge_status = unmerge_fill(ws, issues)
        if unmerge_status:
            print(f">> Unmerge Done")
        else:
            print(f">> Unmerge Failed")
            return False

        # Step 2: Remove Unnecessary Node Header
        remove_node_header_status = remove_node_header(ws, issues)
        if remove_node_header_status:
            print(f">> Remove node header done")
        else:
            print(f">> Remove node header failed")
            return False

        # Step 3: Highlighting empty cells in col1, col2, col3 and row1, row2
        highlight_empty_cell_status=highlight_empty_cell(ws,issues)
        if highlight_empty_cell_status:
            print(f">> Empty cell highlight done")
        else:
            print(f">> Empty cell highlight failed")
            return False

        # Step 4: Apply formatting
        align_status = formatting(ws)
        if align_status:
            print(f">> Alignment done")
        else:
            print(f">> Alignment failed")
            return False

        # Step 5: Save processed workbook
        wb.save(output_file_name)
        wb.close()
        return True
    except PermissionError:
        print(f">> Permission denied. Make sure '{input_file_name}' or '{output_file_name}' is not open in Excel")
        return False
    except Exception as e:
        print(f">> Error processing file: {str(e)}")
        return False


if __name__ == "__main__":
    # ======================
    # Script Configuration
    # ======================
    input_file_name = "input.xlsx"                     # Input Excel file
    sheet_name = "Node Version Planner"                # Target sheet
    # file_datetime=datetime.now().strftime("%d-%m-%Y_%H-%M-%S") # Example timestamp format
    # output_file_name= "unmerged_output_" + str(file_datetime) + ".xlsx"
    output_file_name = f"{input_file_name.strip('.xlsx')}_unmerged_output.xlsx"          # Output Excel file
    issues = {'merged_empty_cells': [],
              'empty_cells_after_unmerge':[],
              'removed_header_rows':[],
              'skipped_rows_during_json_creation':[]}                 # Track issues like empty merged cells

    # Run processing for a single file
    status = processing_excel(input_file_name, sheet_name, output_file_name, issues)

    # Final status check
    if status:
        print(f"✅ Successfully processed '{input_file_name}' -> '{output_file_name}'")
        print("")
        print("⚠️ ISSUES:")
        for issue in issues.keys():
            print("●",issue.upper(),":\n",issues[issue])
            print("")
        # Json creation
        json_output_file = f"{output_file_name.strip('.xlsx')}.json"
        json_status = create_hierarchical_json(load_workbook(output_file_name)[sheet_name], json_output_file, issues)
        if json_status:
            print(f"✅ Output JSON created: '{json_output_file}'")
        else:
            print(f"❌ Failed to create JSON file")
    else:
        print(f"❌ Failed to process '{input_file_name}'")
# Project: Node Details Parser
# Module: Post Process
# Description: To create a JSON file for Excel sheet containing node details.
# Author: EIMACAH

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import json
import logging
import argparse
from openpyxl import load_workbook
from lib.utils import get_last_col_with_value, get_last_row_with_value, setup_logging


# -------------------------------
# Core Processing Function
# -------------------------------

def create_hierarchical_json(
    ws,
    json_output_file,
    skipped_rows_for_json,
    start_version="",
    end_version=""
):
    """
    Create hierarchical JSON from worksheet data between specified version columns.
    Structure: tech -> node_type -> node_version -> [supported_versions]
    """
    try:
        logging.info(f"Starting JSON creation for file: {json_output_file}")
        last_row = get_last_row_with_value(ws)
        last_col = get_last_col_with_value(ws)
        data = {}

        # Get version row (row 2) from column 4 to last column
        version_row = [
            "" if v is None else str(v).strip()
            for v in list(
                ws.iter_rows(
                    min_row=2,
                    max_row=2,
                    min_col=4,
                    max_col=last_col - 1,
                    values_only=True
                )
            )[0]
        ]

        # Check if any empty ENM Version cell in Row2
        if None in version_row or "" in version_row:
            print(">> Error: ENM Version Row2 has some 'None' or '' value")
            logging.error(">> Error: ENM Version Row2 has some 'None' or '' value")
            return False

        # Find start column index based on start_version
        if start_version in (None, ""):
            start_col = 4  # Default start from column 4
            logging.debug("Using default start column: 4")
        else:
            if start_version in version_row:
                start_col = version_row.index(start_version) + 4  # +4 because we started from column 4
                logging.debug(f"Start version '{start_version}' found at column {start_col}")
            else:
                print(f">> Start version '{start_version}' not found in Version Row")
                logging.error(f"Start version '{start_version}' not found in Version Row")
                return False

        # Find end column index based on end_version
        if end_version in (None, ""):
            end_col = last_col - 1
            logging.debug(f"Using default end column: {end_col}")
        else:
            if end_version in version_row:
                end_col = version_row.index(end_version) + 4  # +4 for column offset
                logging.debug(f"End version '{end_version}' found at column {end_col}")
            else:
                print(f">> End version '{end_version}' not found in Version Row")
                logging.error(f"End version '{end_version}' not found in Version Row")
                return False

        # Validate column range
        if start_col >= end_col:
            print(">> Error: start_version must be less than end_version")
            logging.error("start_version must be less than end_version")
            return False

        logging.info(f"Processing rows 3 to {last_row}, columns {start_col} to {end_col}")

        # Process each data row (starting from row 3)
        for row in range(3, last_row + 1):
            # Extract tech, node_type, node_version from columns 1, 2, 3
            tech = ws.cell(row=row, column=1).value
            node_type = ws.cell(row=row, column=2).value
            node_version = ws.cell(row=row, column=3).value
            comments = ws.cell(row=row, column=last_col).value

            # Clean and validate values
            tech = str(tech).strip() if tech is not None else ""
            node_type = str(node_type).strip() if node_type is not None else ""
            node_version = str(node_version).strip() if node_version is not None else ""
            comments = str(comments).strip() if comments is not None else ""

            # Skip rows with missing required data
            if (
                tech in ("", None)
                or node_type in ("", None)
                or node_version in ("", None)
                or tech == node_type == node_version
            ):
                logging.warning(
                    f"Skipping incomplete row {row}: "
                    f"tech='{tech}', node_type='{node_type}', node_version='{node_version}'"
                )
                skipped_rows_for_json.append(row)
                continue

            # Extract supported nodes from version columns within range
            supported_nodes = {}
            for col in range(start_col, end_col + 1):
                val = ws.cell(row=row, column=col).value
                if val not in (None, ""):
                    supported_nodes[str(ws.cell(row=2, column=col).value)] = val

            logging.debug(
                f"Row {row}: {tech}/{node_type}/{node_version} supports {len(supported_nodes)} versions"
            )

            # Build nested dictionary structure
            if len(supported_nodes) > 0:
                if comments not in (None, ""):
                    supported_nodes[str(ws.cell(row=2, column=last_col).value)] = comments
                data.setdefault(tech, {}).setdefault(node_type, {})[node_version] = supported_nodes

        # Write JSON to file with proper encoding
        with open(json_output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

        logging.info(f"JSON file successfully created: {json_output_file}")
        return True

    except Exception as e:
        print(f">> Function:{create_hierarchical_json.__name__}, Error: {str(e)}")
        logging.error(f"Function:{create_hierarchical_json.__name__}, Error: {str(e)}")
        return False



# -------------------------------
# Main Entry Function
# -------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel node details into a hierarchical JSON file."
    )

    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to the input Excel file (e.g., node_details.xlsx)"
    )
    parser.add_argument(
        "--sheetname", "-x",
        required=True,
        help="Path to Sheet name for the input Excel file (e.g., 'Node Version Planner')"
    )
    parser.add_argument(
        "--output", "-o",
        required=False,
        help=(
            "Path to the output JSON file. "
            "If not provided, it will be generated automatically."
        )
    )
    parser.add_argument(
        "--start-version", "-s",
        required=False,
        default="",
        help="Optional: Start version to filter columns (default: first column)"
    )
    parser.add_argument(
        "--end-version", "-e",
        required=False,
        default="",
        help="Optional: End version to filter columns (default: last column)"
    )
    parser.add_argument(
        "--log", "-l",
        nargs='?',
        const=True,
        default=False,
        help=(
            "Enable logging to a file. "
            "Optionally provide a filename; if not, it defaults to <output_basename>_log.txt"
        )
    )

    args = parser.parse_args()

    # Validate input file
    if not os.path.exists(args.input):
        print(f">> Error: Input file '{args.input}' does not exist.")
        return False

    # Auto-generate output name if not provided
    if not args.output:
        base_name = os.path.splitext(os.path.basename(args.input))[0]
        args.output = f"{base_name}_output.json"
    else:
        base_name = os.path.splitext(os.path.basename(args.output))[0]
        args.output = f"{base_name}.json"

    # Call logging setup
    setup_logging(base_name, args.log)

    logging.info(">> Starting Post Process Node Details Parser")
    logging.info(f">> Input: {args.input}")
    logging.info(f">> Output: {args.output}")

    wb = load_workbook(args.input)

    if args.sheetname not in wb.sheetnames:
        print(f">> '{args.sheetname}' does not exist. Available sheet names are \"{wb.sheetnames}\"")
        logging.error(f"Sheet '{args.sheetname}' not found. Available sheets: {wb.sheetnames}")
        return False

    ws = wb[args.sheetname]
    skipped_rows = []

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_dir = os.path.join(base_dir, "output")
    os.makedirs(output_dir, exist_ok=True)

    success = create_hierarchical_json(
        ws=ws,
        json_output_file=os.path.join(output_dir,args.output),
        skipped_rows_for_json=skipped_rows,
        start_version=args.start_version,
        end_version=args.end_version
    )

    if success:
        print(f">> JSON created successfully: {args.output}")
        if skipped_rows:
            print(f">> Skipped rows: {skipped_rows}")
        return True
    else:
        print(">> JSON creation failed.")
        return False


# -------------------------------
# Entry Point
# -------------------------------

if __name__ == "__main__":
    main()

# Project: Node Details Parser
# Module: Pre Process
# Description: To process Excel files containing node details.
# Author: EIMACAH

import os
import logging
from openpyxl import load_workbook
from lib.excel_parser import validate_header_count, unmerge_fill, highlight_empty_cell, highlight_unusable_rows ,formatting


# ==============================================================
# Main Excel Processing Function
# ==============================================================

def processing_excel(input_file, sheet_name, output_file, issues):
    """
    Process and normalize Excel sheet:
      1. Validate structure
      2. Unmerge and fill merged cells
      3. Highlight empty cells
      4. Apply formatting
      5. Save processed output
    """
    try:
        logging.info(f"Starting Excel processing: {input_file} -> {output_file}")

        if not os.path.exists(input_file):
            print(f">> '{input_file}' does not exist")
            logging.error("Input file not found")
            return False

        wb = load_workbook(input_file, read_only=False)
        if sheet_name not in wb.sheetnames:
            print(f">> '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            logging.error("Invalid sheet name")
            return False

        ws = wb[sheet_name]
        logging.debug(f"Worksheet '{sheet_name}' selected")

        # Step 0: Validation
        if not validate_header_count(ws):
            print(">> Worksheet validation failed")
            logging.error("Worksheet validation failed")
            return False
        print(">> Worksheet validation passed")
        logging.info("Worksheet validation completed successfully")

        # Step 1: Unmerge
        if not unmerge_fill(ws, issues):
            print(">> Unmerge failed")
            logging.error("Cell unmerging failed")
            return False
        print(">> Unmerge completed")
        logging.info("Cell unmerging completed successfully")

        # Step 2: Highlight empty cells
        if not highlight_empty_cell(ws, issues):
            print(">> Empty cell highlighting failed")
            logging.error("Empty cell highlighting failed")
            return False
        print(">> Empty cell highlighting completed")
        logging.info("Empty cell highlighting completed successfully")

        # Step 3: Highlight unusable / node header rows
        if not highlight_unusable_rows(ws, issues):
            print(">> Unusable / node header row highlighting failed")
            logging.error("Unusable / node header row highlighting failed")
            return False
        print(">> Unusable / node header row highlighting completed")
        logging.info("Unusable / node header row highlighting completed successfully")

        # Step 4: Apply formatting
        if not formatting(ws):
            print(">> Cell formatting failed")
            logging.error("Cell formatting failed")
            return False
        print(">> Cell formatting completed")
        logging.info("Cell formatting completed successfully")

        # Step 5: Save output
        wb.save(output_file)
        wb.close()

        logging.info(f"Processing completed successfully: {output_file}")
        return True

    except PermissionError:
        print(">> Permission denied. Close the Excel file if open.")
        logging.error("Permission error while accessing files")
        return False

    except Exception as e:
        print(f">> Error processing file: {str(e)}")
        logging.error(f"Unhandled error: {str(e)}")
        return False

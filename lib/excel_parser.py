# Project: Node Details Parser
# Module: Excel Parser
# Description: Functions to normalize Excel sheet containing node details.
# Author: EIMACAH

import logging
from openpyxl.styles import Alignment, PatternFill
from lib.utils import get_last_col_with_value, get_last_row_with_value

# ================================
# Cell Fill Colors
# ================================
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
light_red_fill = PatternFill(start_color="ff6666", end_color="ff6666", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


# ==============================================================
# Excel Processing Utility Functions
# ==============================================================

def formatting(ws):
    """Apply center alignment and text wrapping to all non-empty cells."""
    try:
        logging.info("Starting cell formatting")

        # Define center alignment style with text wrapping
        center_align = Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=True
        )

        # Apply alignment to all non-empty cells
        cell_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = center_align
                    cell_count += 1

        logging.info(f"Cell formatting completed. Formatted {cell_count} cells")
        return True

    except Exception as e:
        print(f">> Function:{formatting.__name__}, Error processing file: {str(e)}")
        logging.error(f"Function:{formatting.__name__}, Error processing file: {str(e)}")
        return False


def remove_node_header(ws, issues):
    """Remove rows where all three main columns have identical values or empty node_type/node_version."""
    try:
        logging.info("Starting node header removal")
        last_row = get_last_row_with_value(ws)
        for index in range(3, last_row + 1):
            cell1 = ws.cell(row=index, column=1).value  # tech
            cell2 = ws.cell(row=index, column=2).value  # node_type
            cell3 = ws.cell(row=index, column=3).value  # node_version

            # Delete row if all three values are identical or if node_type/node_version are empty
            if (cell1 == cell2 == cell3) or (cell2 in (None, "") and cell3 in (None, "")):
                issues['removed_header_rows'].append(index)
                logging.debug(
                    f"Removing header row {index}: tech='{cell1}', node_type='{cell2}', node_version='{cell3}'"
                )
                ws.delete_rows(index)

        logging.info(f"Node header removal completed. Removed {len(issues['removed_header_rows'])} rows")
        return True

    except Exception as e:
        print(f">> Function:{remove_node_header.__name__}, Error processing file: {str(e)}")
        logging.error(f"Function:{remove_node_header.__name__}, Error processing file: {str(e)}")
        return False

def highlight_unusable_rows(ws, issues):
    """
    Highlight rows based on values in Col1, Col2, Col3:
    1. If all three are empty -> highlight full row (light red) [unusable_rows].
    2. If any filled but not all -> highlight full row (light red) + mark missing ones (dark red) [incomplete_rows].
    3. If all three are filled with the same value -> highlight full row (light red) [node_header_rows].
    Updates the issues dict and returns True on success, False on error.
    """
    unusable_rows = []
    incomplete_rows = []
    node_header_rows = []

    try:
        last_col = get_last_col_with_value(ws)
        last_row = get_last_row_with_value(ws)

        # Start from row 3 to skip headers
        for row in range(3, last_row + 1):
            try:
                col1 = ws.cell(row=row, column=1).value
                col2 = ws.cell(row=row, column=2).value
                col3 = ws.cell(row=row, column=3).value

                # Normalize values (strip spaces if string)
                values = [str(v).strip() if isinstance(v, str) else v for v in (col1, col2, col3)]

                if all(v in (None, "") for v in values):
                    # Case 1: all empty
                    unusable_rows.append(row)
                    for col in range(1, last_col + 1):
                        ws.cell(row=row, column=col).fill = light_red_fill
                    logging.debug(f"Row {row} unusable -> highlighted full row (light red)")

                elif values[0] not in (None, "") and values[0] == values[1] == values[2]:
                    # Case 3: all same value -> header row
                    node_header_rows.append(row)
                    for col in range(1, last_col + 1):
                        ws.cell(row=row, column=col).fill = light_red_fill
                    logging.debug(f"Row {row} node header -> highlighted full row (light red)")

                elif any(v in (None, "") for v in values):
                    # Case 2: partially empty -> highlight row and missing cells
                    incomplete_rows.append(row)
                    for col in range(1, last_col + 1):
                        ws.cell(row=row, column=col).fill = light_red_fill
                    for i, v in enumerate(values):
                        if v in (None, ""):
                            ws.cell(row=row, column=i + 1).fill = red_fill
                    logging.debug(f"Row {row} partial -> row filled light red, empty Cols red")

                else:
                    # All three filled and different -> do nothing
                    continue

            except Exception as e_row:
                logging.error(f"Error processing row {row}: {e_row}")

        # Update issues dictionary
        issues['unusable_rows'] = unusable_rows
        issues['incomplete_rows'] = incomplete_rows
        issues['node_header_rows'] = node_header_rows

        return True

    except Exception as e:
        logging.error(f"Error in highlight_unusable_rows: {e}")
        return False


def highlight_empty_cell(ws, issues):
    """Highlight empty cells in critical columns/rows with red fill."""
    try:
        logging.info("Starting empty cell highlighting")
        last_row = get_last_row_with_value(ws)

        # Check rows 1, 2 (headers) in version columns for empty cells
        last_col = get_last_col_with_value(ws)
        for row in (2,):
            for col in range(4, last_col + 1):
                val = ws.cell(row=row, column=col).value
                if val is None or str(val).strip() == "":
                    ws.cell(row=row, column=col).fill = red_fill
                    issues['empty_cell_in_enm_version_row2'].append(ws.cell(row=row, column=col).coordinate)
                    logging.debug(
                        f"Highlighted empty ENM Version Row2 cell at {ws.cell(row=row, column=col).coordinate}"
                    )

        # # Check columns 1, 2, 3 (tech, node_type, node_version) for empty cells
        # for col in (1, 2, 3):
        #     for row in range(3, last_row + 1):
        #         val = ws.cell(row=row, column=col).value
        #         if val is None or str(val).strip() == "":
        #             ws.cell(row=row, column=col).fill = red_fill
        #             issues['empty_cells_after_unmerge'].append(ws.cell(row=row, column=col).coordinate)
        #             logging.debug(f"Highlighted empty cell at {ws.cell(row=row, column=col).coordinate}")

        # total = len(issues['empty_cells_after_unmerge']) + len(issues['empty_cell_in_enm_version_row2'])
        total = len(issues['empty_cell_in_enm_version_row2'])
        logging.info(f"Empty cell highlighting completed. Highlighted {total} cells")
        return True

    except Exception as e:
        print(f">> Function:{highlight_empty_cell.__name__}, Error processing file: {str(e)}")
        logging.error(f"Function:{highlight_empty_cell.__name__}, Error processing file: {str(e)}")
        return False


def unmerge_and_fill(ws, issues):
    """Unmerge all merged cells and fill with original value or highlight if empty."""
    try:
        logging.info("Starting cell unmerging")

        merged_count = len(list(ws.merged_cells.ranges))
        logging.debug(f"Found {merged_count} merged cell ranges")

        # Iterate over all merged cell ranges (create list copy to avoid modification during iteration)
        for merge_range in list(ws.merged_cells.ranges):
            # Extract the top-left cell value of the merged range
            cell_value = ws[merge_range.coord.split(":")[0]].value
            if cell_value is None:
                issues['merged_empty_cells'].append(merge_range)
                logging.debug(f"Empty merged cell range found: {merge_range}")

            # Unmerge the cell range
            ws.unmerge_cells(str(merge_range))
            logging.debug(f"Unmerged range: {merge_range}")

            # Fill each cell in the unmerged range
            for row in ws[merge_range.coord]:
                for cell in row:
                    if cell_value is None:
                        cell.fill = red_fill  # Highlight empty merged cells
                    else:
                        cell.value = cell_value  # Copy merged value to all cells

        logging.info(f"Cell unmerging completed. Processed {merged_count} merged ranges")
        return True

    except Exception as e:
        print(f">> Function:{unmerge_and_fill.__name__}, Error processing file: {str(e)}")
        logging.error(f"Function:{unmerge_and_fill.__name__}, Error processing file: {str(e)}")
        return False


def validate_header_count(ws):
    """Validate worksheet has minimum required columns and rows for processing."""
    try:
        logging.info("Starting worksheet validation")

        last_col = get_last_col_with_value(ws)
        last_row = get_last_row_with_value(ws)
        min_required_columns = 5  # 3 fixed columns + 1 supported node + 1 comment
        min_required_rows = 3     # Header rows + at least 1 data row

        logging.debug(f"Worksheet dimensions: {last_col} columns, {last_row} rows")

        if last_col < min_required_columns:
            print(f">> Sheet has only {last_col} columns. Expected at least {min_required_columns}.")
            logging.error(f"Sheet has only {last_col} columns. Expected at least {min_required_columns}.")
            return False

        if last_row < min_required_rows:
            print(f">> Sheet has only {last_row} rows. Expected at least {min_required_rows}.")
            logging.error(f"Sheet has only {last_row} rows. Expected at least {min_required_rows}.")
            return False

        print(f">> Sheet has {last_col} columns, {last_row} rows — structure looks valid.")
        logging.info(f"Worksheet validation passed: {last_col} columns, {last_row} rows")
        return True

    except Exception as e:
        print(f">> Function:{validate_header_count.__name__}, Error: {str(e)}")
        logging.error(f"Function:{validate_header_count.__name__}, Error: {str(e)}")
        return False

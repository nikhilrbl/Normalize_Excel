# Project: Normalize Excel
# Description: Script to process Excel sheets by unmerging merged cells,
#              filling values, applying formatting, and saving the result.
# Author: EHCIKNA

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime
import os
import json
import argparse
import logging

# ================================
# Color Definitions for Cell Fills
# ================================
# These colors can be applied to highlight issues or mark cells.
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Blue
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Pink
# Extra optional colors
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Purple
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray


def create_hierarchical_json(ws, json_output_file, issues, start_version="", end_version=""):
    """
    Create hierarchical JSON from worksheet data between specified version columns.
    Structure: tech -> nodetype -> nodename -> [supported_versions]
    """
    try:
        logging.info(f"Starting JSON creation for file: {json_output_file}")
        last_row = get_last_row_with_value(ws)
        last_col = get_last_col_with_value(ws)
        # Get version row (row 2) from column 4 to last column
        version_row = [
            "" if v is None else str(v).strip()
            for v in list(ws.iter_rows(
                min_row=2, max_row=2,
                min_col=4, max_col=last_col - 1,
                values_only=True
            ))[0]
        ]
        data = {}
        # print(version_row)
        # logging.debug(f"Version row extracted: {version_row}")

        # Find start column index based on start_version
        if start_version in (None, ""):
            start_col = 4  # Default start from column 4
            logging.debug("Using default start column: 4")
        else:
            if start_version in version_row:
                start_col = version_row.index(start_version) + 4  # +4 because we started from column 4
                logging.debug(f"Start version '{start_version}' found at column {start_col}")
            else:
                print(f">> Start version '{start_version}' not found in Version Row")
                logging.error(f"Start version '{start_version}' not found in Version Row")
                return False

        # Find end column index based on end_version
        if end_version in (None, ""):
            end_col = last_col + 1  # +1 for range() exclusivity
            logging.debug(f"Using default end column: {end_col}")
        else:
            if end_version in version_row:
                end_col = version_row.index(end_version) + 4 + 1  # +4 for column offset, +1 for range() exclusivity
                logging.debug(f"End version '{end_version}' found at column {end_col}")
            else:
                print(f">> End version '{end_version}' not found in Version Row")
                logging.error(f"End version '{end_version}' not found in Version Row")
                return False

        # Validate column range
        if start_col >= end_col:
            print(">> Error: start_version must be less than end_version")
            logging.error("start_version must be less than end_version")
            return False

        logging.info(f"Processing rows {3} to {last_row}, columns {start_col} to {end_col - 1}")

        # Process each data row (starting from row 3)
        for row in range(3, last_row + 1):
            # Extract tech, nodetype, nodename from columns 1, 2, 3
            tech = ws.cell(row=row, column=1).value
            nodetype = ws.cell(row=row, column=2).value
            nodename = ws.cell(row=row, column=3).value

            # Clean and validate values
            tech = str(tech).strip() if tech is not None else ""
            nodetype = str(nodetype).strip() if nodetype is not None else ""
            nodename = str(nodename).strip() if nodename is not None else ""

            # Skip rows with missing required data
            if not tech or not nodetype or not nodename:
                issues['skipped_rows_during_json_creation'].append(row)
                logging.warning(
                    f"Skipping incomplete row {row}: tech='{tech}', nodetype='{nodetype}', nodename='{nodename}'")
                continue  # Skip incomplete rows

            # Extract supported nodes from version columns within range
            supported_nodes = []
            for col in range(start_col, end_col):  # exclude last column
                val = ws.cell(row=row, column=col).value
                if val not in (None, ""):
                    supported_nodes.append(str(ws.cell(row=2, column=col).value))

            logging.debug(f"Row {row}: {tech}/{nodetype}/{nodename} supports {len(supported_nodes)} versions")

            # Build nested dictionary structure
            data.setdefault(tech, {}).setdefault(nodetype, {})[nodename] = supported_nodes

        # Write JSON to file with proper encoding
        with open(json_output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

        logging.info(f"JSON file successfully created: {json_output_file}")
        return True
    except Exception as e:
        print(f">> Function:{create_hierarchical_json.__name__}, Error: {str(e)}")
        logging.error(f"Function:{create_hierarchical_json.__name__}, Error: {str(e)}")
        return False


def get_last_row_with_value(ws):
    """Find the last row containing any non-empty value."""
    for row in range(ws.max_row, 0, -1):  # Iterate backwards from max row
        if any(cell.value not in (None, "") for cell in ws[row]):
            logging.debug(f"Last row with value: {row}")
            return row
    logging.debug("No rows with values found")
    return 0


def get_last_col_with_value(ws):
    """Find the last column containing any non-empty value."""
    for col in range(ws.max_column, 0, -1):  # Iterate backwards from max column
        if any(cell.value not in (None, "") for cell in ws[col]):
            logging.debug(f"Last column with value: {col}")
            return col
    logging.debug("No columns with values found")
    return 0


def remove_node_header(ws, issues):
    """Remove rows where all three main columns have identical values or empty nodetype/nodename."""
    logging.info("Starting node header removal")
    last_row = get_last_row_with_value(ws)
    for index in range(3, last_row + 1):
        cell1 = ws.cell(row=index, column=1).value  # tech
        cell2 = ws.cell(row=index, column=2).value  # nodetype
        cell3 = ws.cell(row=index, column=3).value  # nodename

        # Delete row if all three values are identical or if nodetype/nodename are empty
        if (cell1 == cell2 == cell3) or (cell2 in (None, "") and cell3 in (None, "")):
            issues['removed_header_rows'].append(index)
            logging.debug(f"Removing header row {index}: tech='{cell1}', nodetype='{cell2}', nodename='{cell3}'")
            ws.delete_rows(index)
    logging.info(f"Node header removal completed. Removed {len(issues['removed_header_rows'])} rows")
    return True


def highlight_empty_cell(ws, issues):
    """Highlight empty cells in critical columns/rows with red fill."""
    logging.info("Starting empty cell highlighting")
    last_row = get_last_row_with_value(ws)

    # Check columns 1, 2, 3 (tech, nodetype, nodename) for empty cells
    for col in (1, 2, 3):
        for row in range(3, last_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None or str(val).strip() == "":
                ws.cell(row=row, column=col).fill = red_fill
                issues['empty_cells_after_unmerge'].append(ws.cell(row=row, column=col).coordinate)
                logging.debug(f"Highlighted empty cell at {ws.cell(row=row, column=col).coordinate}")

    # Check rows 1, 2 (headers) in version columns for empty cells
    last_col = get_last_col_with_value(ws)
    for row in (1, 2):
        for col in range(4, last_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is None or str(val).strip() == "":
                ws.cell(row=row, column=col).fill = red_fill
                issues['empty_cells_after_unmerge'].append(ws.cell(row=row, column=col).coordinate)
                logging.debug(f"Highlighted empty header cell at {ws.cell(row=row, column=col).coordinate}")

    logging.info(f"Empty cell highlighting completed. Highlighted {len(issues['empty_cells_after_unmerge'])} cells")
    return True


def formatting(ws):
    """Apply center alignment and text wrapping to all non-empty cells."""
    try:
        logging.info("Starting cell formatting")
        # Define center alignment style with text wrapping
        center_align = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)

        # Apply alignment to all non-empty cells
        cell_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = center_align
                    cell_count += 1

        logging.info(f"Cell formatting completed. Formatted {cell_count} cells")
        return True
    except Exception as e:
        print(f">> Function:{formatting.__name__}, Error processing file: {str(e)}")
        logging.error(f"Function:{formatting.__name__}, Error processing file: {str(e)}")
        return False


def unmerge_fill(ws, issues):
    """Unmerge all merged cells and fill with original value or highlight if empty."""
    try:
        logging.info("Starting cell unmerging")
        merged_count = len(list(ws.merged_cells.ranges))
        logging.debug(f"Found {merged_count} merged cell ranges")

        # Iterate over all merged cell ranges (create list copy to avoid modification during iteration)
        for merge_range in list(ws.merged_cells.ranges):
            # Extract the top-left cell value of the merged range
            cell_value = ws[merge_range.coord.split(":")[0]].value
            if cell_value is None:
                issues['merged_empty_cells'].append(merge_range)
                logging.debug(f"Empty merged cell range found: {merge_range}")

            # Unmerge the cell range
            ws.unmerge_cells(str(merge_range))
            logging.debug(f"Unmerged range: {merge_range}")

            # Fill each cell in the unmerged range
            for row in ws[merge_range.coord]:
                for cell in row:
                    if cell_value is None:
                        cell.fill = red_fill  # Highlight empty merged cells
                    else:
                        cell.value = cell_value  # Copy merged value to all cells

        logging.info(f"Cell unmerging completed. Processed {merged_count} merged ranges")
        return True
    except Exception as e:
        print(f">> Function:{unmerge_fill.__name__}, Error processing file: {str(e)}")
        logging.error(f"Function:{unmerge_fill.__name__}, Error processing file: {str(e)}")
        return False


def validate_header_count(ws):
    """Validate worksheet has minimum required columns and rows for processing."""
    try:
        logging.info("Starting worksheet validation")
        last_col = get_last_col_with_value(ws)
        last_row = get_last_row_with_value(ws)
        min_required_columns = 5  # 3 fixed columns + 1 supported node + 1 comment
        min_required_rows = 3  # Header rows + at least 1 data row

        logging.debug(f"Worksheet dimensions: {last_col} columns, {last_row} rows")

        if last_col < min_required_columns:
            print(f">> Sheet has only {last_col} columns. Expected at least {min_required_columns}.")
            logging.error(f"Sheet has only {last_col} columns. Expected at least {min_required_columns}.")
            return False

        if last_row < min_required_rows:
            print(f">> Sheet has only {last_row} rows. Expected at least {min_required_rows}.")
            logging.error(f"Sheet has only {last_row} rows. Expected at least {min_required_rows}.")
            return False

        print(f">> Sheet has {last_col} columns, {last_row} rows — structure looks valid.")
        logging.info(f"Worksheet validation passed: {last_col} columns, {last_row} rows")
        return True
    except Exception as e:
        print(f">> Function:{validate_header_count.__name__}, Error: {str(e)}")
        logging.error(f"Function:{validate_header_count.__name__}, Error: {str(e)}")
        return False


def processing_excel(*file_info):
    """
    Main processing function to normalize Excel sheet:
    1. Validate structure 2. Unmerge cells 3. Remove headers 4. Highlight issues 5. Format cells
    """
    input_file_name = file_info[0]
    sheet_name = file_info[1]
    output_file_name = file_info[2]
    issues = file_info[3]

    try:
        logging.info(f"Starting Excel processing: {input_file_name} -> {output_file_name}")

        # Check input file existence
        if not os.path.exists(input_file_name):
            print(f">> '{input_file_name}' does not exist")
            logging.error(f"Input file does not exist: {input_file_name}")
            return False

        # Load workbook for processing
        wb = load_workbook(input_file_name, read_only=False)
        logging.debug(f"Workbook loaded successfully")

        # Validate sheet exists in workbook
        if sheet_name not in wb.sheetnames:
            print(f">> '{sheet_name}' does not exist. Available sheet names are \"{wb.sheetnames}\"")
            logging.error(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            return False

        ws = wb[sheet_name]
        logging.debug(f"Worksheet '{sheet_name}' selected")

        # Step 0: Validation - Check minimum rows and columns in excel sheet
        validate_status = validate_header_count(ws)
        if validate_status:
            print(f">> Node Version Planner sheet validation done")
            logging.info("Worksheet validation completed successfully")
        else:
            print(f">> Node Version Planner sheet validation failed")
            logging.error("Worksheet validation failed")
            return False

        # Step 1: Unmerge merged cells and fill with values
        unmerge_status = unmerge_fill(ws, issues)
        if unmerge_status:
            print(f">> Unmerge done")
            logging.info("Cell unmerging completed successfully")
        else:
            print(f">> Unmerge Failed")
            logging.error("Cell unmerging failed")
            return False

        # Step 2: Remove unnecessary node header rows
        remove_node_header_status = remove_node_header(ws, issues)
        if remove_node_header_status:
            print(f">> Remove node header done")
            logging.info("Node header removal completed successfully")
        else:
            print(f">> Remove node header failed")
            logging.error("Node header removal failed")
            return False

        # Step 3: Highlight empty cells in critical columns and header rows
        highlight_empty_cell_status = highlight_empty_cell(ws, issues)
        if highlight_empty_cell_status:
            print(f">> Empty cell highlight done")
            logging.info("Empty cell highlighting completed successfully")
        else:
            print(f">> Empty cell highlight failed")
            logging.error("Empty cell highlighting failed")
            return False

        # Step 4: Apply center alignment and text wrapping formatting
        align_status = formatting(ws)
        if align_status:
            print(f">> Alignment done")
            logging.info("Cell formatting completed successfully")
        else:
            print(f">> Alignment failed")
            logging.error("Cell formatting failed")
            return False

        # Step 5: Save processed workbook to output file
        wb.save(output_file_name)
        wb.close()
        logging.info(f"Excel processing completed successfully: {output_file_name}")
        return True
    except PermissionError:
        print(f">> Permission denied. Make sure '{input_file_name}' or '{output_file_name}' is not open in Excel")
        logging.error(f"Permission denied accessing files: {input_file_name} or {output_file_name}")
        return False
    except Exception as e:
        print(f">> Error processing file: {str(e)}")
        logging.error(f"Error processing file: {str(e)}")
        return False


if __name__ == "__main__":
    # ======================
    # Argument Parsing
    # ======================
    parser = argparse.ArgumentParser(
        description="Normalize Excel sheets by unmerging merged cells, filling values, applying formatting, and saving the result.")
    parser.add_argument("-i", "--input_file", help="Input Excel file path (required)")
    parser.add_argument("-x", "--sheet_name", help="Target worksheet name (required)")
    parser.add_argument("-o", "--output",
                        help="Output Excel file path (optional, defaults to input_unmerged_output.xlsx)")
    parser.add_argument("-s", "--start-version", default="",
                        help="Start version for JSON creation (optional, defaults to beginning)")
    parser.add_argument("-e", "--end-version", default="",
                        help="End version for JSON creation (optional, defaults to end)")
    parser.add_argument("-j", "--json", nargs="?", const=True,
                        help="Export to JSON (optional). Use -j for default filename or -j filename.json for custom filename")
    parser.add_argument("-l", "--logging", nargs="?", const=True,
                        help="Enable logging (optional). Use -l for default log filename or -l filename.log for custom log filename")

    args = parser.parse_args()

    # ======================
    # Validate Required Arguments
    # ======================
    if not args.input_file:
        parser.error("Input file is required. Use -i or --input_file to specify the Excel file path.")

    if not args.sheet_name:
        parser.error("Sheet name is required. Use -x or --sheet_name to specify the worksheet name.")

    # ======================
    # Logging Configuration
    # ======================
    if args.logging:
        # Create Logs directory if it doesn't exist
        logs_dir = "Logs"
        os.makedirs(logs_dir, exist_ok=True)

        # Determine log filename
        if args.logging is True:
            # Default log filename based on output file with date and time
            temp_output_file_name = args.output if args.output else f"{args.input_file.strip('.xlsx')}_normalized_output.xlsx"
            file_datetime = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            log_filename = os.path.join(logs_dir, f"{temp_output_file_name.strip('.xlsx')}_{file_datetime}.log")
        else:
            # Custom log filename provided
            log_filename = os.path.join(logs_dir, args.logging)

        # Configure logging with only file output (no console output for both default and custom)
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8')
            ]
        )

        logging.info(f"Logging enabled. Log file: {log_filename}")
    else:
        # Disable logging if not requested
        logging.disable(logging.CRITICAL)

    # ======================
    # Script Configuration
    # ======================
    input_file_name = args.input_file  # Input Excel file path from argument
    sheet_name = args.sheet_name  # Target worksheet name from argument
    # file_datetime=datetime.now().strftime("%d-%m-%Y_%H-%M-%S") # Example timestamp format
    # output_file_name= "unmerged_output_" + str(file_datetime) + ".xlsx"

    # Create Output directory if it doesn't exist
    output_dir = "Output"
    os.makedirs(output_dir, exist_ok=True)

    # Set output file path in Output directory
    if args.output:
        output_file_name = os.path.join(output_dir, os.path.basename(args.output))
    else:
        output_file_name = os.path.join(output_dir,
                                        f"{os.path.basename(input_file_name).strip('.xlsx')}_normalized_output.xlsx")

    issues = {'merged_empty_cells': [],  # Track empty merged cell ranges
              'empty_cells_after_unmerge': [],  # Track empty cells found during processing
              'removed_header_rows': [],  # Track removed header row indices
              'skipped_rows_during_json_creation': []}  # Track skipped rows during JSON creation

    # Run main processing pipeline for the Excel file
    status = processing_excel(input_file_name, sheet_name, output_file_name, issues)

    # Final status check and reporting
    if status:
        print(f"✅ Successfully processed '{input_file_name}' -> '{output_file_name}'")
        print("")
        print("⚠️ ISSUES:")
        for issue in issues.keys():
            print("●", issue.upper(), ":\n", issues[issue])
            print("")

        # Create hierarchical JSON from processed Excel data (only if -j flag is provided)
        if args.json:
            if args.json is True:
                # Default JSON filename in Output directory
                json_output_file = os.path.join(output_dir, f"{os.path.basename(output_file_name).strip('.xlsx')}.json")
            else:
                # Custom JSON filename provided in Output directory
                json_output_file = os.path.join(output_dir, os.path.basename(args.json))

            json_status = create_hierarchical_json(load_workbook(output_file_name)[sheet_name], json_output_file,
                                                   issues, args.start_version, args.end_version)
            if json_status:
                print(f"✅ Output JSON created: '{json_output_file}'")
            else:
                print(f"❌ Failed to create JSON file")
    else:
        print(f"❌ Failed to process '{input_file_name}'")